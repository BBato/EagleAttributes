import csv
import openpyxl
import digiKeyInterface
import farnell_interface
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from tkinter import messagebox


def execute_main(inputFileName, inputMasterEPN, optGetFarnellCodes, optGenerateSpecSheets, optGenerateEagleScript, optGenerateBOM, progressbar, printToConsole):


    try:
        # Prepare input file for reading CSV data
        with open(inputFileName) as f:
            reader = csv.reader(f, delimiter=';')
            data = list(reader)
    except:
        # The specified file cannot be read
        return "Cannot open file: "+inputFileName

    # Method to parse CSV and determine column header locations
    def col(header):
        for x in range(0, len(data[0])-1):
            if data[0][x] == header:
                return x

        # If header does not exist return -1
        return -1

    # Match header names and save them as location values
    colParts = col("Parts")
    colDigikey = col("DIGIKEY_PARTNUM")
    colQty = col("Qty")
    colValue = col("Value")
    colDevice = col("Device")
    colDesc = col("Description")
    colPackage = col("Package")
    colMf = col("MANUFACTURER_NAME")
    colEPN = col("EPN")
    colSupplier = col("SUPPLIER")

    # The DigiKey and EPN columns are required as minimum
    if colDigikey == -1:
        return "[ERROR] No attribute 'DIGIKEY_PARTNUM' found. Add this attribute to some of your components in Eagle."

    if colEPN == -1:
        return "[ERROR] No attribute 'EPN' found. Add this attribute to some of your components in Eagle."    

    # Initialize statistics for assmebly table
    numThroughHoleComponents = 0
    numSurfaceMountComponents = 0
    numUniqueSurfaceMountComponents = 0
    numUniqueParts = 0

    # Prepare file for writing EAGLE attribute commands
    if optGenerateEagleScript:
        eagleFile = open("output/eagle.txt", "w")

    # Prepare file for writing output BOM
    if optGenerateBOM:
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Append one line above the headers, and then a line with headers
        sheet.append(["BOM"])
        sheet.append(['EPN','Qty','Value','Parts','Description','DIGIKEY_PARTNUM','FARNELL_OC','MANUFACTURER','MANUFACTURER_PARTNUM','SUPPLIER','TECH'])

    # Repeat for each entry in EAGLE CSV file
    for x in range(1,len(data)):

        # Notify user of progress via progressbar
        progressbar["value"]=(x/len(data))*100
        progressbar.update()

        # There can be multiple components per row, split by comma
        components = str(data[x][colParts]).split(",")

        # Each row owns one DigiKey code
        digiKeyCode = str(data[x][colDigikey])
        
        # No need to print the below components to BOM
        if "MOUNTING HOLE" in data[x][colDesc] or "MOUNTING PAD" in data[x][colDesc] or "TEST PIN" in data[x][colDesc]:
            printToConsole("[WARNING] Skipping element "+data[x][colDesc])
            continue

        # Placing an asterix in a component value indicates the engineer wants a new resistor with the specified value
        if "*" in data[x][colValue]:

            # -- Need to find an alternative resistor --
            alternativeResistorRequired = True

            # Get new resistor based on specified value
            digiKeyCode = digiKeyInterface.getAlternativeResistor(data[x][colValue].replace("*",""), data[x][colPackage],printToConsole)

        else:
            
            # If no DigiKey code is available, fill in cells directly from CSV and continue.
            alternativeResistorRequired = False

            # DigiKey code should not be empty. If it is, forward data from EAGLE CSV to BOM directly and quit this component
            if digiKeyCode == "":
                if optGenerateBOM:
                    sheet.append([data[x][colEPN], data[x][colQty], data[x][colValue], data[x][colParts], data[x][colDesc], "-", "-", data[x][colMf], "-", "-", "-"])

                continue

        # Parse the DigiKey website and get relevant data on this component
        digiKeyData = digiKeyInterface.getDigiKeyReference(digiKeyCode, optGenerateSpecSheets, printToConsole, data[x][colEPN], alternativeResistorRequired )
        farnellCode = "Farnell"

        # If we are changing the resistor, remember to update its value and EPN in CSV too
        if alternativeResistorRequired:
            data[x][colValue] = digiKeyData[5]
            data[x][colEPN] = digiKeyData[6]

        # In all cases overwrite the supplier column
        data[x][colSupplier] = digiKeyData[7]

        # Limit the use of Farnell API by asking the user to select this opton first
        if optGetFarnellCodes:
            farnellCode = farnell_interface.getFarnell(digiKeyData[1], printToConsole)

        # Init this variable
        duplicateComponent = False

        # If the component being parsed is not the first on the CSV
        if x>1:

            # Parse through all previously saved components in the BOM
            for y in range(3,sheet.max_row+1):

                # Get each Digikey code and compare to the incoming code
                digiKeyCode2 = sheet["F"+str(y)].value
                if digiKeyCode2 == digiKeyCode:
                    
                    # If the codes are the same, we need to merge the rows
                    duplicateComponent = True
                    printToConsole("[WARNING] Duplicate component "+sheet["F"+str(y)].value)

                    # Merge parts
                    newParts = sheet["D"+str(y)].value

                    # Add quantities
                    newQty = int(sheet["B"+str(y)].value)
                    newQty += len(components)

                    # Select an EPN from one of the two parts, whichever has one
                    newEPN = sheet["A"+str(y)].value
                    if newEPN == "":
                        newEPN = data[x][colEPN]

                    # Merge components
                    for c in components:
                        newParts += ", "+c.strip()
                    
                    # Write these new values to the existing row in BOM
                    sheet.cell(row = y, column = 1, value = newEPN)
                    sheet.cell(row = y, column = 2, value = str(newQty))
                    sheet.cell(row = y, column = 4, value = newParts)

                    break

        # Proceed to write a new row in the BOM if the component wasn't merged with an existing row
        if optGenerateBOM:
            if not duplicateComponent:
                sheet.append([data[x][colEPN], data[x][colQty], data[x][colValue], data[x][colParts], digiKeyData[3], digiKeyCode, "-", digiKeyData[0], digiKeyData[1], data[x][colSupplier], digiKeyData[4]])
                numUniqueParts = numUniqueParts + 1

                # Add hyperlinks to the BOM
                sheet.cell(row=sheet.max_row, column=6).hyperlink = digiKeyData[7]
                sheet.cell(row=sheet.max_row, column=10, value="DigiKey").hyperlink = "https://www.digikey.cn/products/zh?WT.z_header=search_go&keywords="+digiKeyCode

                if optGetFarnellCodes:
                    if farnellCode is not "-":
                        sheet.cell(row=sheet.max_row, column=7, value = farnellCode).hyperlink = "https://ie.farnell.com/search?st="+farnellCode

        # Gather statistics on throughhole components and surface mount components
        if digiKeyData[4] == "TH":
            numThroughHoleComponents = numThroughHoleComponents + len(components)
        if digiKeyData[4] == "SMD":
            numSurfaceMountComponents = numSurfaceMountComponents + len(components)
            
            if not duplicateComponent:
                numUniqueSurfaceMountComponents = numUniqueSurfaceMountComponents + 1

        # Draft an EAGLE script which will be used to update component attributes in the schematic
        if optGenerateEagleScript:

            for c in components:
                
                partName = c.strip()
                eagleFile.write("ATTRIBUTE "+partName+" MANUFACTURER_NAME '"+digiKeyData[0]+"'; \n")
                eagleFile.write("CHANGE DISPLAY OFF; \n")
                eagleFile.write("ATTRIBUTE "+partName+" MANUFACTURER_PART_NUMBER '"+digiKeyData[1]+"'; \n")
                eagleFile.write("CHANGE DISPLAY OFF; \n")
                eagleFile.write("ATTRIBUTE "+partName+" DK_AVAILABILITY '"+digiKeyData[2]+"'; \n")
                eagleFile.write("CHANGE DISPLAY OFF; \n")
                eagleFile.write("ATTRIBUTE "+partName+" DK_DESCRIPTION '"+digiKeyData[3]+"'; \n")
                eagleFile.write("CHANGE DISPLAY OFF; \n")

                if optGetFarnellCodes:
                    eagleFile.write("ATTRIBUTE "+partName+" OC_FARNELL '"+farnellCode+"'; \n")
                    eagleFile.write("CHANGE DISPLAY OFF; \n")

                if alternativeResistorRequired:
                    eagleFile.write("VALUE "+partName+" '"+digiKeyData[5]+"'; \n")
                    eagleFile.write("ATTRIBUTE "+partName+" DIGIKEY_PARTNUM '"+digiKeyCode+"'; \n")
                    eagleFile.write("ATTRIBUTE "+partName+" Description '"+digiKeyData[3]+"'; \n")

    if optGenerateEagleScript:
        eagleFile.close()

    if optGenerateBOM:

        # Adjust BOM sheet columns width per expected content
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 5
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 50
        sheet.column_dimensions['F'].width = 20
        sheet.column_dimensions['G'].width = 10
        sheet.column_dimensions['H'].width = 20
        sheet.column_dimensions['I'].width = 20
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 10

        #Adjust sheet font (top left)
        fontStyle = Font(size = "18")
        sheet.cell(row = 1, column = 1, value = inputMasterEPN+' BOM').font = fontStyle

        # define main table style
        mediumStyle = openpyxl.worksheet.table.TableStyleInfo(name='TableStyleMedium4',
                                                            showRowStripes=True)

        # create main table
        table = openpyxl.worksheet.table.Table(ref="A2:K"+str(sheet.max_row),
                                            displayName='BOM',
                                            tableStyleInfo=mediumStyle)
        # add the table to the worksheet
        sheet.add_table(table)

        # merge main table header
        sheet.merge_cells("A1:K1")

        assemby_table_row = sheet.max_row + 2

        # create assembly info table header
        sheet.cell(row = assemby_table_row, column = 1, value = 'Assembly Info').font = fontStyle

        sheet.cell(row = assemby_table_row+1, column = 1, value = 'Assembly Options')
        sheet.cell(row = assemby_table_row+2, column = 1, value = 'Unique Part Count')
        sheet.cell(row = assemby_table_row+3, column = 1, value = 'SMD Unique Part Count')
        sheet.cell(row = assemby_table_row+4, column = 1, value = 'SMD Total Part Count')
        sheet.cell(row = assemby_table_row+5, column = 1, value = 'Through Hole Part Count')

        sheet.cell(row = assemby_table_row+2, column = 4, value = numUniqueParts)
        sheet.cell(row = assemby_table_row+3, column = 4, value = numUniqueSurfaceMountComponents)
        sheet.cell(row = assemby_table_row+4, column = 4, value = numSurfaceMountComponents)
        sheet.cell(row = assemby_table_row+5, column = 4, value = numThroughHoleComponents)

        # Give a special black border to the assembly table
        def allRoundBorder():
            return Border(top = Side(border_style='thin', color='FF000000'),    
                              right = Side(border_style='thin', color='FF000000'), 
                              bottom = Side(border_style='thin', color='FF000000'),
                              left = Side(border_style='thin', color='FF000000'))

        for y in range(0,6):
            for x in range(1,5):
                sheet.cell(row=assemby_table_row+y, column=x).border = allRoundBorder() 

        # merge assembly info table header
        sheet.merge_cells("A"+str(assemby_table_row)+":D"+str(assemby_table_row))

        # merge assembly info table body
        sheet.merge_cells("A"+str(assemby_table_row+1)+":C"+str(assemby_table_row+1))
        sheet.merge_cells("A"+str(assemby_table_row+2)+":C"+str(assemby_table_row+2))
        sheet.merge_cells("A"+str(assemby_table_row+3)+":C"+str(assemby_table_row+3))
        sheet.merge_cells("A"+str(assemby_table_row+4)+":C"+str(assemby_table_row+4))
        sheet.merge_cells("A"+str(assemby_table_row+5)+":C"+str(assemby_table_row+5))

        # Try to save, but if another application is blocking, notify the user to close it and try last time.
        try:
            wb.save("output/"+inputMasterEPN+".xlsx")
        except:
            messagebox.showinfo("Error","Saving failed because the file is open in an external editor. Close it and press OK to try again.")
            wb.save("output/"+inputMasterEPN+".xlsx")
        

    return "Complete."


